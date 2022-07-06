import json
import datetime
from threading import Thread
import os
import sys

sys.path.append('..')

from emoji import emojize
import openpyxl
from django.http import JsonResponse, FileResponse, Http404

from database_new import DATABASE
from uuid_token import forge_token, get_uuid_from_token
from controller import CONTROLLER
from email_sender import EmailSender
from run_search import SEARCH_RUNNER


def search(request):
    """搜索请求处理函数

    url: 42.192.44.52:8000/search/

    解包前端发送的json，将各字段组合成搜索语句，将搜索记录保存至数据库

    **启动搜索服务**：在这里添加搜索服务的入口
    """
    # 解包前端请求
    if request.method == 'GET':
        if CONTROLLER.emoji_status is True:
            print(emojize(':white_check_mark: 已收到 search 请求', language='alias'))
            print(request.GET)

        token = request.GET.get('token')
        keywords = request.GET.get('keywords')
        start_time = request.GET.get('start_time')
        end_time = request.GET.get('end_time')
        article_type = request.GET.get('article_type')
        language = request.GET.get('language')
        species = request.GET.get('species')
        sex = request.GET.get('sex')
        age = request.GET.get('age')

    if request.method == 'POST':
        if CONTROLLER.emoji_status is True:
            print(emojize(':white_check_mark: 已收到 search 请求', language='alias'))
            print(request.POST)

        token = request.POST.get('token')
        keywords = request.POST.get('keywords')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        article_type = request.POST.get('article_type')
        language = request.POST.get('language')
        species = request.POST.get('species')
        sex = request.POST.get('sex')
        age = request.POST.get('age')

        if CONTROLLER.emoji_status is True:
            print("article_type: ", article_type)
            print("language: ", language)
            print("species: ", species)
            print("sex: ", sex)
            print("age: ", age)

    # 从token中获取uuid
    uuid_str = get_uuid_from_token(token)
    # token时效性判断
    if uuid_str == 'token expired':
        json_rsp = {"message_type": "token_expired"}
    else:
        uuid = int(uuid_str)

        # 拼接搜索语句
        if start_time is None:
            start_time = '1900-01-01'
        if end_time is None:
            end_time = datetime.datetime.now().strftime('%Y-%m-%d')
        robust_keywords = '(' + keywords + ') AND ("' + start_time + '"[Date - Publication]:' + '"' + end_time + '"[Date - Publication])'
        article_type_str = ''
        if article_type is not None and len(article_type) > 0 and article_type != '[]':
            article_type = article_type.replace('[', '').replace('"', '').replace(']', '').split(',')
            if len(article_type) > 0:
                robust_keywords += ' AND ('
                for f in article_type:
                    robust_keywords += '(' + f + '[FILT]) OR ('
                    article_type_str += f + ' OR '
                # article_type_int = ['0', '0', '0', '0', '0', '0']
                # for f in article_type:
                #     robust_keywords += '(' + f + '[FILT]) OR ('
                #     if f == 'Books and Documents':
                #         article_type_int[0] = '1'
                #     if f == 'Clinical Trial':
                #         article_type_int[1] = '1'
                #     if f == 'Meta-Analysis':
                #         article_type_int[2] = '1'
                #     if f == 'Randomized Controlled Trial':
                #         article_type_int[3] = '1'
                #     if f == 'Review':
                #         article_type_int[4] = '1'
                #     if f == 'Systematic Review':
                #         article_type_int[5] = '1'
                # cache = ''
                # for i in article_type_int:
                #     cache += i
                # article_type_int = int(cache)
                robust_keywords = robust_keywords[:-5] + ')'
                article_type_str = article_type_str[:-4]
        else:
            article_type_str = ''
        language_str = ''
        if language is not None and len(language) > 0 and language != '[]':
            language = language.replace('[', '').replace('"', '').replace(']', '').split(',')
            if len(language) > 0:
                robust_keywords += ' AND ('
                for f in language:
                    robust_keywords += '(' + f + '[Language]) OR ('
                    language_str += f + ','
                language_str = language_str[:-1]
                robust_keywords = robust_keywords[:-5] + ')'
        else:
            language = ''
        species_str = ''
        if species is not None and len(species) > 0 and species != '[]':
            species = species.replace('[', '').replace('"', '').replace(']', '').split(',')
            if len(species) > 0:
                robust_keywords += ' AND ('
                for f in species:
                    robust_keywords += '(' + f + '[FILT]) OR ('
                    species_str += f + ','
                species_str = species_str[:-1]
                robust_keywords = robust_keywords[:-5] + ')'
        else:
            species = ''
        sex_str = ''
        if sex is not None and len(sex) > 0 and sex != '[]':
            sex = sex.replace('[', '').replace('"', '').replace(']', '').split(',')
            if len(sex) > 0:
                robust_keywords += ' AND ('
                for f in sex:
                    robust_keywords += '(' + f + '[FILT]) OR ('
                    sex_str += f + ','
                sex_str = species_str[:-1]
                robust_keywords = robust_keywords[:-5] + ')'
        else:
            sex = ''
        age_str = ''
        if age is not None and len(age) > 0 and age != '[]':
            age = age.replace('[', '').replace('"', '').replace(']', '').split(',')
            if len(age) > 0:
                robust_keywords += ' AND ('
                for f in age:
                    robust_keywords += '(' + f + '[FILT]) OR ('
                    age_str += f + ','
                age_str = age_str[:-1]
                robust_keywords = robust_keywords[:-5] + ')'
        else:
            age = ''

        print("Search keywords: ", robust_keywords)

        # 保存搜索记录
        print(uuid)
        print(keywords)
        print(start_time)
        print(end_time)
        print(article_type_str)
        print(age_str)
        print(language_str)
        print(species_str)
        print(sex_str)
        timestamp = DATABASE.add_search_history(uuid, keywords, start_time, end_time, article_type_str, age_str, language_str, species_str, sex_str)

        # 开启一个单独的线程运行搜索服务并在搜索完成后执行善后处理
        # search_thread = Thread(target=run_search, args=(robust_keywords, timestamp, keywords, start_time, end_time, article_type_int, age, language, species, sex))
        # search_thread.start()

        # 将搜索任务放入任务队列
        SEARCH_RUNNER.search_task_queue.append((robust_keywords, timestamp, keywords, start_time, end_time, article_type_str, age, language, species, sex))

        # 向前端返回响应
        new_token = forge_token(uuid_str)
        json_rsp = {
            'message_type': 'search_received',
            'token': new_token
        }

    if CONTROLLER.emoji_status is True:
        print(emojize(':rocket: 已发送 search_received 应答', language='alias'))
        print(json_rsp)

    cache = JsonResponse(json_rsp)
    cache["Access-Control-Allow-Origin"] = "*"
    return cache


# def run_search(robust_keywords: str, timestamp: int, raw_keywords, start_time, end_time, filter_article_type, filter_age, filter_language, filter_species, filter_sex):
#     """适用于单个线程的执行搜索及搜索完成善后任务的函数
#
#     将 robust_keywords 输入搜索脚本，启动搜索服务；完成搜索后将本次搜索在 search_history 表中的条目修改为已完成搜索
#
#     :param robust_keywords: (str) 经过鲁棒性处理后的搜索关键词组合
#     :param timestamp: (int) 本次搜索对应的搜索记录的时间戳
#     :return: None
#     """
#     # 获取结果保存路径
#     history = DATABASE.get_result(timestamp)[0]
#     result_timestamp = history[1]
#     result_dir = 'static/search_result/' + str(result_timestamp) + '/'
#
#     # TODO: 启动搜索，使用 robust_keywords 作为关键词进行搜索，将 paper_info.xlsx 和 clue_info.xlsx 两个文件输出到 result_dir 下
#
#     # 从json文件中提取结果文献的 abstract_highlight 字段，生成HTML标签，保存至数据库 paper_abstract 表
#     if CONTROLLER.test_mode is True:
#         # 更改至测试用路径
#         result_dir = 'static/json_data_test/'
#
#     # 读取 file.list 获取包含结果文献信息的json文件目录
#     file_list = open(result_dir + 'file.list', 'r')
#     json_dir_list = file_list.readlines()
#     for json_dir in json_dir_list:
#         json_dir = json_dir[:-1]    # 删除换行符 \n
#         raw_paper_info = json.load(open(json_dir, 'r'))
#         pmid = int(raw_paper_info['pmid'])
#         abstract_highlight_list = raw_paper_info['abstract_highlight']
#         # 生成带有高亮的HTML字符串
#         abstract_highlight_str = ''
#         for i in range(len(abstract_highlight_list)):
#             word = abstract_highlight_list[i]
#             if word != '':
#                 if word[0] == '<':
#                     color_sp = word.find('font_color=') + 11
#                     color_ep = word.find('>')
#                     color = word[color_sp:color_ep]
#                     abstract_highlight_str += '<span style="font-weight: bold; font-style: italic; color: ' + color + ';">'
#                     abstract_highlight_str += abstract_highlight_list[i + 1] + '</span>'
#                     try:
#                         abstract_highlight_list[i + 1] = ''
#                     except:
#                         pass
#                 else:
#                     abstract_highlight_str += word
#         # 写入数据库
#         DATABASE.add_paper_highlight_abstract(pmid, abstract_highlight_str)
#
#     # 将数据库中所有搜索该关键词的搜索记录标记为已完成搜索
#     DATABASE.search_completed(raw_keywords, start_time, end_time, filter_article_type, filter_age, filter_language, filter_species, filter_sex)
#
#     # 向发起搜索的用户发送提醒邮件
#     # 生成验证码
#     uuids = DATABASE.get_uuids_with_keywords(robust_keywords)
#     for uuid in uuids:
#         uuid = uuid[0]
#         user = DATABASE.get_user(uuid)
#         email = user.email
#         username = user.username
#         email_sender = EmailSender(email, username)
#         email_sender.generate_search_completed_content(robust_keywords)
#         # 发送验证邮件
#         email_sender.send()


def get_history(request):
    """获取用户历史记录请求处理函数

    url: 42.192.44.52:8000/search/history/

    使用uuid在数据库中查找该用户的历史记录，以json形式发送到前端
    """
    # 解包前端请求
    if request.method == 'GET':
        token = request.GET.get('token')
    if request.method == 'POST':
        token = request.POST.get('token')

    if CONTROLLER.emoji_status is True:
        print(emojize(':white_check_mark: 已收到 get_history 请求', language='alias'))
        print(emojize(':snake: token: ' + token, language='alias'))

    # 从token中获取uuid
    uuid_str = get_uuid_from_token(token)
    # token时效性判断
    if uuid_str == 'token expired':
        json_rsp = {"message_type": "token_expired"}
    else:
        uuid = int(uuid_str)
        # 使用uuid查找用户的历史记录。注意，这里查询到的不一定是全部记录，这取决于 Database.get_search_history 中设定的最大条数。
        history_list = DATABASE.get_search_history(uuid)
        history_list = history_list[::-1]
        # print(history_list)
        new_token = forge_token(uuid_str)
        json_rsp = {
            'message_type': 'history_list',
            'history': [],
            'token': new_token
        }
        # 将历史记录放入 'history' 字段
        for history in history_list:
            timestamp = history[0]
            raw_keywords = history[2]
            start_time = history[6]
            end_time = history[7]
            # article_type_int = str(history[8])
            # article_type = ''
            # for i in range(len(article_type_int)):
            #     if article_type_int[i] == '1':
            #         if i == 0:
            #             article_type += 'Books and Documents'
            #         elif i == 1:
            #             if article_type != '':
            #                 article_type += 'OR Clinical Trial'
            #             else:
            #                 article_type += 'Clinical Trial'
            #         elif i == 2:
            #             if article_type != '':
            #                 article_type += 'OR Meta-Analysis'
            #             else:
            #                 article_type += 'Meta-Analysis'
            #         elif i == 3:
            #             if article_type != '':
            #                 article_type += 'OR Randomized Controlled Trial'
            #             else:
            #                 article_type += 'Randomized Controlled Trial'
            #         elif i == 4:
            #             if article_type != '':
            #                 article_type += 'OR Review'
            #             else:
            #                 article_type += 'Review'
            #         elif i == 5:
            #             if article_type != '':
            #                 article_type += 'OR Systematic Review'
            #             else:
            #                 article_type += 'Systematic Review'
            # if article_type != '':
            #     article_type = '(' + article_type + '[FILT])'
            article_type = history[8]
            age = history[9]
            if age.find(',') != -1:
                age_ls = age.split(',')
                age = '(' + age_ls[0] + ' OR ' + age_ls[1] + '[FILT])'
            else:
                if age != '':
                    age = '(' + age + '[FILT])'
            language = history[10]
            if language.find(',') != -1:
                language_ls = language.split(',')
                language = '(' + language_ls[0] + ' OR ' + language_ls[1] + '[FILT])'
            else:
                if language != '':
                    language = '(' + language + '[FILT])'
            species = history[11]
            if species.find(',') != -1:
                species_ls = species.split(',')
                species = '(' + species_ls[0] + ' OR ' + species_ls[1] + '[FILT])'
            else:
                if species != '':
                    species = '(' + species + '[FILT])'
            sex = history[12]
            if sex.find(',') != -1:
                sex_ls = sex.split(',')
                sex = '(' + sex_ls[0] + ' OR ' + sex_ls[1] + '[FILT])'
            else:
                if sex != '':
                    sex = '(' + sex + '[FILT])'

            robust_keywords = '(' + raw_keywords + ') AND ("' + start_time + '"[Date - Publication]:' + '"' + \
                              end_time + '"[Date - Publication])'
            if article_type != '':
                robust_keywords += ' AND ' + article_type
            if age != '':
                robust_keywords += ' AND ' + age
            if language != '':
                robust_keywords += 'AND' + language
            if species != '':
                robust_keywords += ' AND ' + species
            if sex != '':
                robust_keywords += ' AND ' + sex

            history_dic = {'timestamp': timestamp, 'raw_keywords': raw_keywords, 'robust_keywords': robust_keywords}
            json_rsp['history'].append(history_dic)

    if CONTROLLER.emoji_status is True:
        print(emojize(':rocket: 已发送 rsp_get_history 应答', language='alias'))
        print(json_rsp)

    cache = JsonResponse(json_rsp)
    cache["Access-Control-Allow-Origin"] = "*"
    return cache


# def get_result(request):
#     """获取搜索结果请求处理函数
#
#     url: 42.192.44.52:8000/search/result/
#
#     使用时间戳在数据库中查找搜索结果（result_timestamp），然后到 /static/search_result/result_timestamp/ 中读取excel文件，
#     最后形成json发送至前端
#     """
#     # 解包前端请求
#     if request.method == 'GET':
#         token = request.GET.get('token')
#         timestamp = request.GET.get('timestamp')
#         page_num = request.GET.get('page_num')
#     if request.method == 'POST':
#         token = request.POST.get('token')
#         timestamp = request.POST.get('timestamp')
#         page_num = request.POST.get('page_num')
#
#     # 请求合法性判断
#     # 从token中获取uuid
#     uuid_str = get_uuid_from_token(token)
#     # token时效性判断
#     if uuid_str == 'token expired':
#         json_rsp = {"message_type": "token_expired"}
#         cache = JsonResponse(json_rsp)
#         cache["Access-Control-Allow-Origin"] = "*"
#         return cache
#     else:
#         print("Loading data from excel files!")
#         uuid = int(uuid_str)
#         timestamp = int(timestamp)
#         # 使用 timestamp 到数据库中查找此条历史记录
#         history = DATABASE.get_result(timestamp)[0]
#         if uuid == int(history[4]):     # 身份验证通过
#             new_token = forge_token(uuid_str)
#             json_rsp = {
#                 "message_type": "result",
#                 "paper_info": [],
#                 "clue_info": [],
#                 "token": new_token
#             }
#             result_timestamp = history[1]
#             file_dir = 'static/search_result/' + str(result_timestamp) + '/'
#
#             # 读取 'paper_info.xlsx'，将信息放入 "paper_info" 字段
#             file_name = file_dir + 'paper_info.xlsx'
#             workbook = openpyxl.load_workbook(file_name)
#             worksheet = workbook.get_sheet_by_name("Sheet1")
#             row_max = worksheet.max_row
#             page_num = int(page_num)
#
#             # 每页显示的文章数目
#             papers_on_one_page = 10
#
#             row_start = (page_num - 1) * 10 + 2
#             row_end = row_start + papers_on_one_page
#
#             if row_start < row_max:
#                 if row_end > row_max:
#                     row_end = row_max
#                 for i in range(row_start, row_end):
#                     paper_info = {}
#                     paper_info['Pmid'] = worksheet.cell(i, 1).value
#                     paper_info['Journal'] = worksheet.cell(i, 2).value
#                     paper_info['Publication_Type'] = worksheet.cell(i, 3).value
#                     paper_info['Publication_Year'] = worksheet.cell(i, 4).value
#                     paper_info['Publication_Date'] = worksheet.cell(i, 5).value
#                     paper_info['Title'] = worksheet.cell(i, 6).value
#                     paper_info['First_Author'] = worksheet.cell(i, 7).value
#                     paper_info['Corresponding_Author'] = worksheet.cell(i, 8).value
#                     paper_info['Authors'] = worksheet.cell(i, 9).value
#                     paper_info['Affiliations'] = worksheet.cell(i, 10).value
#                     paper_info['Abstract'] = worksheet.cell(i, 11).value
#                     paper_info['Keywords'] = worksheet.cell(i, 12).value
#                     paper_info['Doi'] = worksheet.cell(i, 13).value
#                     paper_info['Journal_If'] = worksheet.cell(i, 14).value
#                     paper_info['Conclusion'] = worksheet.cell(i, 15).value
#                     paper_info['Chinese_Title'] = worksheet.cell(i, 16).value
#                     paper_info['Chinese_Abstract'] = worksheet.cell(i, 17).value
#                     paper_info['Sample_Size'] = worksheet.cell(i, 18).value
#                     paper_info['Location'] = worksheet.cell(i, 19).value
#                     paper_info['Organization'] = worksheet.cell(i, 20).value
#
#                     json_rsp["paper_info"].append(paper_info)
#
#             # 读取 'clue_info.xlsx'，将信息放入 "clue_info" 字段
#             file_name = file_dir + 'clue_info.xlsx'
#             workbook = openpyxl.load_workbook(file_name)
#             worksheet = workbook.get_sheet_by_name("Sheet1")
#             row_max = worksheet.max_row
#
#             for i in range(2, row_max):
#                 clue_info = {}
#                 clue_info['Node1'] = worksheet.cell(i, 1).value
#                 clue_info['Edge_Type'] = worksheet.cell(i, 2).value
#                 clue_info['Node2'] = worksheet.cell(i, 3).value
#                 clue_info['Weight'] = worksheet.cell(i, 4).value
#                 clue_info['Paper_List'] = worksheet.cell(i, 5).value
#                 clue_info['Original_Text'] = worksheet.cell(i, 6).value
#
#                 json_rsp["clue_info"].append(clue_info)
#
#             cache = JsonResponse(json_rsp)
#             cache["Access-Control-Allow-Origin"] = "*"
#             return cache
#
#         else:   # 发送请求的用户与历史记录所属用户不匹配
#             json_rsp = {"message_type": "invalid_request"}
#             cache = JsonResponse(json_rsp)
#             cache["Access-Control-Allow-Origin"] = "*"
#             return cache


def get_paper_info(request):
    """获取文献列表请求处理函数

    url: 42.192.44.52:8000/search/paper_info/

    使用时间戳在数据库中查找搜索结果（result_timestamp），然后到 /static/search_result/result_timestamp/ 中读取excel文件，
    最后形成json发送至前端
    """
    # 解包前端请求
    if request.method == 'GET':
        token = request.GET.get('token')
        timestamp = request.GET.get('timestamp')
        page_num = request.GET.get('page_num')
        column = request.GET.get('column')
        order = request.GET.get('order')
    if request.method == 'POST':
        token = request.POST.get('token')
        timestamp = request.POST.get('timestamp')
        page_num = request.POST.get('page_num')
        column = request.POST.get('column')
        order = request.POST.get('order')

    if CONTROLLER.emoji_status is True:
        print(emojize(':white_check_mark: 已收到 get_paper_info 请求', language='alias'))
        print(emojize(':snake: token: ' + token, language='alias'))
        print(emojize(':snake: timestamp: ' + timestamp, language='alias'))
        print(emojize(':snake: page_num: ' + page_num, language='alias'))

    # 请求合法性判断
    # 从token中获取uuid
    uuid_str = get_uuid_from_token(token)
    # token时效性判断
    if uuid_str == 'token expired':
        json_rsp = {"message_type": "token_expired"}
    else:
        uuid = int(uuid_str)
        timestamp = int(timestamp)
        # 使用 timestamp 到数据库中查找此条历史记录
        history = DATABASE.get_result(timestamp)[0]
        if uuid == int(history[4]):  # 身份验证通过
            new_token = forge_token(uuid_str)
            json_rsp = {
                "message_type": "paper_info",
                "paper_info": [],
                "token": new_token
            }

            # 获取结果数据的后端目录
            result_timestamp = history[1]
            file_dir = 'static/search_result/' + str(result_timestamp) + '/'

            # 每页显示的文章数目
            papers_on_one_page = 10

            # 使用excel文件作为数据源加载结果数据
            if CONTROLLER.paper_info_data_source == 'excel':
                print("Loading data from excel files...")

                # 读取 'paper_info.xlsx'，将信息放入 "paper_info" 字段
                file_name = file_dir + 'paper_info.xlsx'
                workbook = openpyxl.load_workbook(file_name, read_only=True)
                worksheet = workbook["Sheet1"]
                row_max = worksheet.max_row
                page_num = int(page_num)

                json_rsp['total'] = row_max - 1

                row_start = (page_num - 1) * papers_on_one_page + 2
                row_end = row_start + papers_on_one_page

                if row_start <= row_max:
                    if row_end > row_max:
                        row_end = row_max + 1
                    for i in range(row_start, row_end):
                        paper_info = {}
                        paper_info['Pmid'] = worksheet.cell(i, 1).value
                        paper_info['Journal'] = worksheet.cell(i, 2).value
                        paper_info['Publication_Type'] = worksheet.cell(i, 3).value
                        paper_info['Publication_Year'] = worksheet.cell(i, 4).value
                        paper_info['Publication_Date'] = worksheet.cell(i, 5).value
                        paper_info['Title'] = worksheet.cell(i, 6).value
                        paper_info['First_Author'] = worksheet.cell(i, 7).value
                        paper_info['Corresponding_Author'] = worksheet.cell(i, 8).value
                        paper_info['Authors'] = worksheet.cell(i, 9).value
                        paper_info['Affiliations'] = worksheet.cell(i, 10).value
                        paper_info['Abstract'] = worksheet.cell(i, 11).value
                        paper_info['Keywords'] = worksheet.cell(i, 12).value
                        paper_info['Doi'] = worksheet.cell(i, 13).value
                        paper_info['Journal_If'] = worksheet.cell(i, 14).value
                        paper_info['Conclusion'] = worksheet.cell(i, 15).value
                        paper_info['Chinese_Title'] = worksheet.cell(i, 16).value
                        paper_info['Chinese_Abstract'] = worksheet.cell(i, 17).value
                        paper_info['Sample_Size'] = worksheet.cell(i, 18).value
                        paper_info['Location'] = worksheet.cell(i, 19).value
                        paper_info['Organization'] = worksheet.cell(i, 20).value

                        json_rsp["paper_info"].append(paper_info)

            # 使用json文件作为数据源加载结果数据
            if CONTROLLER.paper_info_data_source == 'json':
                print("Loading data from json files...")

                if CONTROLLER.test_mode is True:
                    # 更改至测试用路径
                    file_dir = 'static/json_data_test/'

                # 读取 file.list 获取包含结果文献信息的json文件目录
                file_list = open(file_dir + 'file.list', 'r')
                json_dir_list = file_list.readlines()
                paper_info_list = []

                for json_dir in json_dir_list:
                    json_dir = json_dir[:-1]
                    raw_paper_info = json.load(open(json_dir, 'r'))
                    paper_info_list.append(raw_paper_info)

                # 根据请求中的 column 和 order 字段对列表 json_dir_list 进行排序
                if column == 'Publication_Date':
                    if order == 'positive':
                        sorted_paper_info = sorted(paper_info_list, key=lambda x: x['publication_date'])
                    else:
                        sorted_paper_info = sorted(paper_info_list, key=lambda x: x['publication_date'], reverse=True)
                elif column == 'Journal_If':
                    if order == 'positive':
                        sorted_paper_info = sorted(paper_info_list, key=lambda x: x['journal_if'])
                    else:
                        sorted_paper_info = sorted(paper_info_list, key=lambda x: x['journal_if'], reverse=True)
                else:
                    pass

                # print(json_dir_list)
                row_max = len(json_dir_list)
                page_num = int(page_num)

                json_rsp['total'] = row_max

                row_start = (page_num - 1) * papers_on_one_page
                row_end = row_start + papers_on_one_page
                # print(row_start)
                # print(row_end)

                if row_start <= row_max:
                    if row_end > row_max:
                        row_end = row_max
                    for i in range(row_start, row_end):
                        raw_paper_info = sorted_paper_info[i]
                        paper_info = {}
                        paper_info['Pmid'] = raw_paper_info['pmid']
                        pmid = int(raw_paper_info['pmid'])
                        paper_info['Journal'] = raw_paper_info['journal']
                        paper_info['Publication_Type'] = raw_paper_info['publication_type']
                        paper_info['Publication_Year'] = raw_paper_info['publication_year']
                        paper_info['Publication_Date'] = raw_paper_info['publication_date']
                        paper_info['Title'] = raw_paper_info['title']
                        paper_info['First_Author'] = raw_paper_info['first_author']
                        paper_info['Corresponding_Author'] = raw_paper_info['corresponding_author']
                        paper_info['Authors'] = raw_paper_info['authors']
                        paper_info['Affiliations'] = raw_paper_info['affiliations']
                        paper_info['Abstract'] = DATABASE.get_highlight_abstract_with_pmid(pmid)
                        paper_info['Keywords'] = raw_paper_info['keywords']
                        paper_info['Doi'] = raw_paper_info['doi']
                        paper_info['Journal_If'] = raw_paper_info['journal_if']
                        paper_info['Chinese_Title'] = raw_paper_info['title_zh']
                        paper_info['Chinese_Abstract'] = raw_paper_info['abstract_zh']
                        paper_info['Sample_Size'] = raw_paper_info['sample_size']
                        # 以下3个字段可能在原始json文件中不存在
                        try:
                            paper_info['Conclusion'] = raw_paper_info['conclusion']
                        except:
                            paper_info['Conclusion'] = ''
                        try:
                            paper_info['Location'] = raw_paper_info['location']
                        except:
                            paper_info['Location'] = ''
                        try:
                            paper_info['Organization'] = raw_paper_info['organization']
                        except:
                            paper_info['Organization'] = ''

                        json_rsp["paper_info"].append(paper_info)
                        # print(paper_info)

        else:   # 发送请求的用户与历史记录所属用户不匹配
            json_rsp = {"message_type": "invalid_request"}

    if CONTROLLER.emoji_status is True:
        print(emojize(':rocket: 已发送 rsp_get_paper_info 应答', language='alias'))
        print(json_rsp)

    cache = JsonResponse(json_rsp)
    cache["Access-Control-Allow-Origin"] = "*"
    return cache


def get_clue_info(request):
    """获取线索列表请求处理函数

    url: 42.192.44.52:8000/search/clue_info/

    使用时间戳在数据库中查找搜索结果（result_timestamp），然后到 /static/search_result/result_timestamp/ 中读取excel文件，
    最后形成json发送至前端
    """
    # 解包前端请求
    if request.method == 'GET':
        token = request.GET.get('token')
        timestamp = request.GET.get('timestamp')
        page_num = request.GET.get('page_num')
        column = request.GET.get('column')
        order = request.GET.get('order')
    if request.method == 'POST':
        token = request.POST.get('token')
        timestamp = request.POST.get('timestamp')
        page_num = request.POST.get('page_num')
        column = request.POST.get('column')
        order = request.POST.get('order')

    if CONTROLLER.emoji_status is True:
        print(emojize(':white_check_mark: 已收到 get_clue_info 请求', language='alias'))
        print(emojize(':snake: token: ' + token, language='alias'))
        print(emojize(':snake: timestamp: ' + timestamp, language='alias'))
        print(emojize(':snake: page_num: ' + page_num, language='alias'))

    # 请求合法性判断
    # 从token中获取uuid
    uuid_str = get_uuid_from_token(token)
    # token时效性判断
    if uuid_str == 'token expired':
        json_rsp = {"message_type": "token_expired"}
    else:
        uuid = int(uuid_str)
        timestamp = int(timestamp)
        # 使用 timestamp 到数据库中查找此条历史记录
        history = DATABASE.get_result(timestamp)[0]
        if uuid == int(history[4]):  # 身份验证通过
            new_token = forge_token(uuid_str)
            json_rsp = {
                "message_type": "clue_info",
                "clue_info": [],
                "token": new_token
            }
            result_timestamp = history[1]
            file_dir = 'static/search_result/' + str(result_timestamp) + '/'
            page_num = int(page_num)

            # 每页显示的线索数目
            papers_on_one_page = 20

            if CONTROLLER.clue_info_data_source == 'excel':
                print("Loading data from excel files...")

                # 读取 'clue_info.xlsx'，将信息放入 "clue_info" 字段
                file_name = file_dir + 'clue_info.xlsx'
                workbook = openpyxl.load_workbook(file_name, read_only=True)
                worksheet = workbook.get_sheet_by_name("Sheet1")
                row_max = worksheet.max_row

                json_rsp['total'] = row_max - 1

                # 如果页号为 0 则代表返回所有条目
                if page_num == 0:
                    row_start = 2
                    row_end = row_max
                    json_rsp['message_type'] = 'network'
                else:
                    row_start = (page_num - 1) * papers_on_one_page + 2
                    row_end = row_start + papers_on_one_page

                if row_start <= row_max:
                    if row_end > row_max:
                        row_end = row_max + 1
                    for i in range(row_start, row_end):
                        clue_info = {}
                        clue_info['Node1'] = worksheet.cell(i, 1).value
                        clue_info['Edge_Type'] = worksheet.cell(i, 2).value
                        clue_info['Node2'] = worksheet.cell(i, 3).value
                        clue_info['Weight'] = worksheet.cell(i, 4).value
                        clue_info['Paper_List'] = worksheet.cell(i, 5).value
                        clue_info['Original_Text'] = worksheet.cell(i, 6).value

                        json_rsp["clue_info"].append(clue_info)

            if CONTROLLER.clue_info_data_source == 'json':
                print("Loading data from json file...")

                # 读取json文件
                file_name = file_dir + 'clue_info.json'

                if CONTROLLER.test_mode is True:
                    # 切换至测试用目录
                    file_name = 'static/search_result/114514/clue_info.json'

                clue_info = json.load(open(file_name, 'r'))['network']

                # 对 clue_info 排序
                if column == 'Weight':
                    if order == 'positive':
                        clue_info = sorted(clue_info, key=lambda x: x['Weight'])
                    else:
                        clue_info = sorted(clue_info, key=lambda x: x['Weight'], reverse=True)
                else:
                    pass

                row_max = len(clue_info)
                json_rsp['total'] = row_max

                if page_num == 0:
                    json_rsp['message_type'] = 'network'
                    json_rsp["clue_info"] = clue_info

                else:
                    row_start = (page_num - 1) * papers_on_one_page
                    row_end = row_start + papers_on_one_page
                    if row_start <= row_max:
                        if row_end > row_max:
                            row_end = row_max
                        json_rsp["clue_info"] = clue_info[row_start:row_end]

            if page_num == 0:
                # 统计 edge_type 类型集合
                edge_type_set = json.load(open(file_name, 'r'))['header']['Edge_Type']
                edge_type_list = []
                for edge_type in edge_type_set.keys():
                    if edge_type_set[edge_type]['has_BFS_edge'] is True:
                        edge_type_list.append(edge_type)
                edge_type_list.sort()
                edge_type_list.insert(0, 'BFS')

                json_rsp["edge_type_list"] = edge_type_list

        else:  # 发送请求的用户与历史记录所属用户不匹配
            json_rsp = {"message_type": "invalid_request"}

    if CONTROLLER.emoji_status is True:
        print(emojize(':rocket: 已发送 rsp_get_clue_info 应答', language='alias'))
        print(json_rsp)

    cache = JsonResponse(json_rsp)
    cache["Access-Control-Allow-Origin"] = "*"
    return cache


def get_paper_details(request):
    """处理通过PMID获取单篇文章详细信息的请求

    url: 42.192.44.52:8000/search/paper_details/
    """
    # 解包前端请求
    if request.method == 'GET':
        token = request.GET.get('token')
        timestamp = request.GET.get('timestamp')
        pmid = request.GET.get('pmid')
    if request.method == 'POST':
        token = request.POST.get('token')
        timestamp = request.POST.get('timestamp')
        pmid = request.POST.get('pmid')

    # 从token中获取uuid
    uuid_str = get_uuid_from_token(token)
    # token时效性判断
    if uuid_str == 'token expired':
        json_rsp = {"message_type": "token_expired"}
    else:
        uuid = int(uuid_str)
        timestamp = int(timestamp)
        pmid_str = pmid
        pmid = int(pmid)
        # 使用 timestamp 到数据库中查找此条历史记录
        history = DATABASE.get_result(timestamp)[0]
        if uuid == int(history[4]):  # 身份验证通过
            new_token = forge_token(uuid_str)
            json_rsp = {
                "message_type": "paper_details",
                "Pmid": pmid,
                "token": new_token
            }
            result_timestamp = history[1]
            file_dir = 'static/search_result/' + str(result_timestamp) + '/'

            if CONTROLLER.test_mode is True:
                # 更改至测试用路径
                file_dir = 'static/json_data_test/'

            # 读取 file.list 获取包含结果文献信息的json文件目录
            file_list = open(file_dir + 'file.list', 'r')
            json_dir_list = file_list.readlines()
            for i in range(len(json_dir_list)):
                json_dir = json_dir_list[i][:-1]  # 删除换行符 \n
                raw_paper_info = json.load(open(json_dir, 'r'))
                if raw_paper_info['pmid'] == pmid_str:
                    paper_info = {}
                    paper_info['Pmid'] = raw_paper_info['pmid']
                    paper_info['Journal'] = raw_paper_info['journal']
                    paper_info['Publication_Type'] = raw_paper_info['publication_type']
                    paper_info['Publication_Year'] = raw_paper_info['publication_year']
                    paper_info['Publication_Date'] = raw_paper_info['publication_date']
                    paper_info['Title'] = raw_paper_info['title']
                    paper_info['First_Author'] = raw_paper_info['first_author']
                    paper_info['Corresponding_Author'] = raw_paper_info['corresponding_author']
                    paper_info['Authors'] = raw_paper_info['authors']
                    paper_info['Affiliations'] = raw_paper_info['affiliations']
                    paper_info['Abstract'] = DATABASE.get_highlight_abstract_with_pmid(pmid)
                    paper_info['Keywords'] = raw_paper_info['keywords']
                    paper_info['Doi'] = raw_paper_info['doi']
                    paper_info['Journal_If'] = raw_paper_info['journal_if']
                    paper_info['Chinese_Title'] = raw_paper_info['title_zh']
                    paper_info['Chinese_Abstract'] = raw_paper_info['abstract_zh']
                    paper_info['Sample_Size'] = raw_paper_info['sample_size']
                    # 以下3个字段可能在原始json文件中不存在
                    try:
                        paper_info['Conclusion'] = raw_paper_info['conclusion']
                    except:
                        paper_info['Conclusion'] = ''
                    try:
                        paper_info['Location'] = raw_paper_info['location']
                    except:
                        paper_info['Location'] = ''
                    try:
                        paper_info['Organization'] = raw_paper_info['organization']
                    except:
                        paper_info['Organization'] = ''

                    json_rsp['paper_info'] = paper_info
                    break

            cache = JsonResponse(json_rsp)
            cache["Access-Control-Allow-Origin"] = "*"
            return cache


def download_file(request):
    """处理结果文件下载请求

    url: 42.192.44.52:8000/search/download/
    """
    # 解包前端请求
    if request.method == 'GET':
        token = request.GET.get('token')
        history_id = request.GET.get('timestamp')
        file_name = request.GET.get('file_name')
    if request.method == 'POST':
        token = request.POST.get('token')
        history_id = request.POST.get('timestamp')
        file_name = request.POST.get('file_name')

    print(file_name)

    # 请求合法性判断
    # 从token中获取uuid
    uuid_str = get_uuid_from_token(token)
    # token时效性判断
    if uuid_str == 'token expired':
        raise Http404
    else:
        uuid = int(uuid_str)
        history_id = int(history_id)
        # 使用 timestamp 到数据库中查找此条历史记录
        history = DATABASE.get_result(history_id)[0]
        if uuid == int(history[4]):  # 身份验证通过
            # 获取结果数据的后端目录
            result_timestamp = history[1]
            file_dir = 'static/search_result/' + str(result_timestamp) + '/'

    if CONTROLLER.test_mode is True:
        file_dir = 'static/search_result/114514/'

    file_path = file_dir + file_name
    try:
        response = FileResponse(open(file_path, 'rb'))
        response['content_type'] = "application/octet-stream"
        response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(file_path)
        response["Access-Control-Allow-Origin"] = "*"
        return response
    except Exception:
        raise Http404
