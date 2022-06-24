import openpyxl

file_dir = '../static/search_result/' + str(114514) + '/'
file_name = file_dir + 'paper_info.xlsx'
workbook = openpyxl.load_workbook(file_name)
worksheet = workbook.get_sheet_by_name("Sheet1")
rows, cols = worksheet.max_row, worksheet.max_column

cell = worksheet.cell(1, 1)
print(cell.value)
